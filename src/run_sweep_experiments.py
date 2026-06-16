#!/usr/bin/env python3
"""
Run a GPU-friendly experiment sweep for the TFG heat-diffusion project.

Where to put this file:
    Preferably in src/run_sweep_experiments.py, next to config.py and main.py.
    It also works from the project root if the project has a src/ folder.

Main outputs:
    src/results/sweep_experiments/all_results.csv
    src/results/sweep_experiments/sweep_report.xlsx

Recommended first commands:
    python src/run_sweep_experiments.py --mode smoke
    python src/run_sweep_experiments.py --mode recommended

Full factorial command requested by Axel:
    python src/run_sweep_experiments.py --mode full

The script is resumable: if all_results.csv already exists, completed rows are skipped.
"""

from __future__ import annotations

import argparse
import itertools
import math
import os
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

# -----------------------------------------------------------------------------
# Import project modules whether this script is executed from project root or src.
# -----------------------------------------------------------------------------
THIS_FILE = Path(__file__).resolve()
THIS_DIR = THIS_FILE.parent
if (THIS_DIR / "config.py").exists():
    sys.path.insert(0, str(THIS_DIR))
elif (THIS_DIR / "src" / "config.py").exists():
    sys.path.insert(0, str(THIS_DIR / "src"))
else:
    # Keep the default import error, but make the problem obvious.
    print(
        "ERROR: cannot find config.py. Put this file in src/ or in the project root.",
        file=sys.stderr,
    )

import pandas as pd
import torch
from torch import nn

from config import ALPHA, CONFIG, DT, DX, GRID_SIZE, NUM_TIMESTEPS, TEST_STEPS, TRAIN_STEPS, VAL_STEPS
from data.graph_builder import build_grid_graph
from data.heat_diffusion import (
    add_gaussian_noise,
    build_temporal_pairs,
    select_data_fraction,
    simulate_heat_diffusion,
)
from models.baselines import MLPBaseline
from models.full_gnn import FullGNN
from models.tiny_gnn import TinyGNN
from training.evaluator import (
    evaluate_physics_violation,
    evaluate_prediction_error,
    measure_model_size_bytes,
)
from training.trainer import model_forward, train_one_epoch
from utils.metrics import count_trainable_parameters
from utils.seed import set_seed


# -----------------------------------------------------------------------------
# Experiment defaults
# -----------------------------------------------------------------------------
DEFAULT_EPOCHS = [1, 10, 50, 100, 300]
DEFAULT_SEEDS = list(range(40, 51))
DEFAULT_HIDDEN_DIMS = [32, 16, 8, 4]
DEFAULT_DATA_FRACTIONS = [1.0, 0.5, 0.2, 0.1, 0.05, 0.01]
DEFAULT_NOISE_LEVELS = [0.0, 0.01, 0.05, 0.1, 0.2]

# Chosen on a log scale to see whether the physics term helps a little,
# dominates too much, or gives a sweet spot. 0.0 is already covered by TinyGNN.
DEFAULT_PHYSICS_LAMBDAS = [0.001, 0.01, 0.1, 1.0, 10.0]

MODEL_ORDER = {
    "MLP": 0,
    "FULL_GNN": 1,
    "TINY_GNN": 2,
    "TINY_GNN_PINN": 3,
}
MODEL_SEED_OFFSETS = {
    "MLP": 101,
    "FULL_GNN": 202,
    "TINY_GNN": 303,
    "TINY_GNN_PINN": 303,  # same architecture/init as TinyGNN; only loss changes
}
KEY_COLS = [
    "mode",
    "seed",
    "checkpoint_epoch",
    "hidden_dim",
    "data_fraction",
    "noise_level",
    "model",
    "physics_lambda",
    "learning_rate",
    "dropout",
]


@dataclass(frozen=True)
class PreparedData:
    x_train_in: torch.Tensor
    x_train_tgt: torch.Tensor
    x_val_in: torch.Tensor
    x_val_tgt: torch.Tensor
    x_test_in: torch.Tensor
    x_test_tgt: torch.Tensor
    edge_index: torch.Tensor
    train_pairs: int
    val_pairs: int
    test_pairs: int


@dataclass(frozen=True)
class SweepCombo:
    hidden_dim: int
    data_fraction: float
    noise_level: float


@dataclass(frozen=True)
class ModelSpec:
    name: str
    display_name: str
    is_pinn: bool
    physics_lambda: float


def parse_int_list(value: str) -> list[int]:
    """Parse '1,2,3' or inclusive range '40:50'."""
    value = value.strip()
    if ":" in value:
        start, end = value.split(":", 1)
        return list(range(int(start), int(end) + 1))
    return [int(x.strip()) for x in value.split(",") if x.strip()]


def parse_float_list(value: str) -> list[float]:
    return [float(x.strip()) for x in value.split(",") if x.strip()]


def unique_preserve_order(items: Iterable[Any]) -> list[Any]:
    seen: set[Any] = set()
    result: list[Any] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return result


def resolve_device(device_arg: str) -> str:
    if device_arg == "auto":
        return "cuda" if torch.cuda.is_available() else "cpu"
    if device_arg == "cuda" and not torch.cuda.is_available():
        raise RuntimeError("You asked for --device cuda, but torch.cuda.is_available() is False.")
    return device_arg


def configure_torch(device: str) -> None:
    if device.startswith("cuda"):
        torch.backends.cudnn.benchmark = True
        try:
            torch.set_float32_matmul_precision("high")
        except Exception:
            pass


def make_sweep_combos(
    mode: str,
    hidden_dims: list[int],
    data_fractions: list[float],
    noise_levels: list[float],
) -> list[SweepCombo]:
    """
    full:        all hidden_dim x data_fraction x noise_level combinations.
    recommended: one-factor-at-a-time design, much cheaper and easier to explain.
    smoke:       minimal test to verify that everything works.
    """
    if mode == "smoke":
        return [SweepCombo(hidden_dim=16, data_fraction=1.0, noise_level=0.0)]

    if mode == "full":
        return [
            SweepCombo(hidden_dim=h, data_fraction=f, noise_level=n)
            for h, f, n in itertools.product(hidden_dims, data_fractions, noise_levels)
        ]

    if mode == "recommended":
        anchor_hidden = 16 if 16 in hidden_dims else hidden_dims[0]
        triples: list[tuple[int, float, float]] = []

        # Capacity sweep: does TinyGNN remain competitive as hidden_dim shrinks?
        triples.extend((h, 1.0, 0.0) for h in hidden_dims)

        # Data-efficiency sweep: does graph/physics help with less data?
        triples.extend((anchor_hidden, f, 0.0) for f in data_fractions)

        # Noise robustness sweep: does PINN regularize noisy supervision?
        triples.extend((anchor_hidden, 1.0, n) for n in noise_levels)

        return [SweepCombo(h, f, n) for h, f, n in unique_preserve_order(triples)]

    raise ValueError(f"Unknown mode: {mode}")


def make_model_specs(only_models: list[str], physics_lambdas: list[float]) -> list[ModelSpec]:
    requested = {m.upper() for m in only_models}
    specs: list[ModelSpec] = []

    def enabled(name: str) -> bool:
        return not requested or name in requested

    if enabled("MLP"):
        specs.append(ModelSpec("MLP", "MLP", is_pinn=False, physics_lambda=0.0))
    if enabled("FULL_GNN"):
        specs.append(ModelSpec("FULL_GNN", "FULL_GNN", is_pinn=False, physics_lambda=0.0))
    if enabled("TINY_GNN"):
        specs.append(ModelSpec("TINY_GNN", "TINY_GNN", is_pinn=False, physics_lambda=0.0))
    if enabled("TINY_GNN_PINN"):
        for lam in physics_lambdas:
            specs.append(ModelSpec("TINY_GNN_PINN", "TINY_GNN + PINN", is_pinn=True, physics_lambda=float(lam)))

    if not specs:
        raise ValueError("No models selected. Use names: MLP,FULL_GNN,TINY_GNN,TINY_GNN_PINN")
    return specs


def build_model(spec: ModelSpec, hidden_dim: int, dropout: float) -> nn.Module:
    if spec.name == "MLP":
        return MLPBaseline(hidden_dim=hidden_dim)
    if spec.name == "FULL_GNN":
        return FullGNN(hidden_dim=hidden_dim, dropout=dropout)
    if spec.name in {"TINY_GNN", "TINY_GNN_PINN"}:
        return TinyGNN(hidden_dim=hidden_dim)
    raise ValueError(f"Unknown model: {spec.name}")


def prepare_data(
    *,
    seed: int,
    data_fraction: float,
    noise_level: float,
    device: str,
) -> PreparedData:
    set_seed(seed)

    sim = simulate_heat_diffusion(
        grid_size=GRID_SIZE,
        num_timesteps=NUM_TIMESTEPS,
        train_steps=TRAIN_STEPS,
        val_steps=VAL_STEPS,
        test_steps=TEST_STEPS,
        alpha=ALPHA,
        dt=DT,
        dx=DX,
        seed=seed,
    )

    x_train_in, x_train_tgt = build_temporal_pairs(sim.train_states)
    x_val_in, x_val_tgt = build_temporal_pairs(sim.val_states)
    x_test_in, x_test_tgt = build_temporal_pairs(sim.test_states)

    x_train_in, x_train_tgt = select_data_fraction(
        x_train_in,
        x_train_tgt,
        fraction=data_fraction,
        seed=seed,
    )

    # Add noise only to training supervision. Validation/test remain clean so
    # noise experiments measure robustness/generalisation, not noisy evaluation.
    if noise_level > 0.0:
        x_train_in = add_gaussian_noise(x_train_in, sigma=noise_level, seed=seed + 17)
        x_train_tgt = add_gaussian_noise(x_train_tgt, sigma=noise_level, seed=seed + 31)

    x_train_in = x_train_in.unsqueeze(-1).to(device)
    x_train_tgt = x_train_tgt.unsqueeze(-1).to(device)
    x_val_in = x_val_in.unsqueeze(-1).to(device)
    x_val_tgt = x_val_tgt.unsqueeze(-1).to(device)
    x_test_in = x_test_in.unsqueeze(-1).to(device)
    x_test_tgt = x_test_tgt.unsqueeze(-1).to(device)

    graph = build_grid_graph(grid_size=GRID_SIZE)
    edge_index = graph.edge_index.to(device)

    return PreparedData(
        x_train_in=x_train_in,
        x_train_tgt=x_train_tgt,
        x_val_in=x_val_in,
        x_val_tgt=x_val_tgt,
        x_test_in=x_test_in,
        x_test_tgt=x_test_tgt,
        edge_index=edge_index,
        train_pairs=int(x_train_in.shape[0]),
        val_pairs=int(x_val_in.shape[0]),
        test_pairs=int(x_test_in.shape[0]),
    )


def cuda_sync_if_needed(device: str) -> None:
    if device.startswith("cuda") and torch.cuda.is_available():
        torch.cuda.synchronize()


def measure_inference_time_ms(
    model: nn.Module,
    x_input: torch.Tensor,
    edge_index: torch.Tensor,
    device: str,
    repeats: int,
) -> float:
    model.eval()
    repeats = max(1, repeats)

    with torch.no_grad():
        for _ in range(min(5, repeats)):
            _ = model_forward(model, x_input[0], edge_index)
        cuda_sync_if_needed(device)
        start = time.perf_counter()
        for _ in range(repeats):
            _ = model_forward(model, x_input[0], edge_index)
        cuda_sync_if_needed(device)
        end = time.perf_counter()

    return ((end - start) / repeats) * 1000.0


def make_row_key(row: dict[str, Any]) -> tuple[Any, ...]:
    return tuple(row[col] for col in KEY_COLS)


def make_expected_key(
    *,
    mode: str,
    seed: int,
    checkpoint_epoch: int,
    combo: SweepCombo,
    spec: ModelSpec,
    learning_rate: float,
    dropout: float,
) -> tuple[Any, ...]:
    return (
        mode,
        seed,
        checkpoint_epoch,
        combo.hidden_dim,
        combo.data_fraction,
        combo.noise_level,
        spec.name,
        spec.physics_lambda,
        learning_rate,
        dropout,
    )


def load_completed_keys(csv_path: Path) -> set[tuple[Any, ...]]:
    if not csv_path.exists():
        return set()
    try:
        df = pd.read_csv(csv_path)
    except Exception:
        return set()
    if df.empty or not set(KEY_COLS).issubset(df.columns):
        return set()
    return {tuple(row[col] for col in KEY_COLS) for _, row in df[KEY_COLS].iterrows()}


def append_rows_to_csv(rows: list[dict[str, Any]], csv_path: Path) -> None:
    if not rows:
        return
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows)
    write_header = not csv_path.exists()
    df.to_csv(csv_path, mode="a", header=write_header, index=False)


def evaluate_checkpoint(
    *,
    mode: str,
    model: nn.Module,
    spec: ModelSpec,
    data: PreparedData,
    seed: int,
    checkpoint_epoch: int,
    max_epochs: int,
    combo: SweepCombo,
    learning_rate: float,
    dropout: float,
    device: str,
    losses: dict[str, float],
    train_seconds_so_far: float,
    inference_repeats: int,
) -> dict[str, Any]:
    cuda_sync_if_needed(device)

    val_mse = evaluate_prediction_error(model, data.x_val_in, data.x_val_tgt, data.edge_index)
    test_mse = evaluate_prediction_error(model, data.x_test_in, data.x_test_tgt, data.edge_index)
    phys_viol = evaluate_physics_violation(
        model,
        data.x_test_in,
        data.edge_index,
        alpha=ALPHA,
        dt=DT,
        dx=DX,
        grid_size=GRID_SIZE,
    )
    inference_ms = measure_inference_time_ms(
        model,
        data.x_test_in,
        data.edge_index,
        device=device,
        repeats=inference_repeats,
    )
    num_params = count_trainable_parameters(model)
    model_size_bytes = measure_model_size_bytes(model)

    row = {
        "run_label": (
            f"{spec.display_name} | {checkpoint_epoch} epoch | seed {seed} | "
            f"hidden {combo.hidden_dim} | data {combo.data_fraction} | "
            f"noise {combo.noise_level} | lambda {spec.physics_lambda}"
        ),
        "mode": mode,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "model": spec.name,
        "model_display": spec.display_name,
        "model_order": MODEL_ORDER[spec.name],
        "is_pinn": bool(spec.is_pinn),
        "seed": seed,
        "checkpoint_epoch": int(checkpoint_epoch),
        "max_epochs": int(max_epochs),
        "hidden_dim": int(combo.hidden_dim),
        "data_fraction": float(combo.data_fraction),
        "noise_level": float(combo.noise_level),
        "physics_lambda": float(spec.physics_lambda),
        "learning_rate": float(learning_rate),
        "dropout": float(dropout if spec.name == "FULL_GNN" else 0.0),
        "device": device,
        "grid_size": GRID_SIZE,
        "num_timesteps": NUM_TIMESTEPS,
        "train_steps": TRAIN_STEPS,
        "val_steps": VAL_STEPS,
        "test_steps": TEST_STEPS,
        "train_pairs": data.train_pairs,
        "val_pairs": data.val_pairs,
        "test_pairs": data.test_pairs,
        "alpha": ALPHA,
        "dt": DT,
        "dx": DX,
        "train_data_loss": float(losses.get("data_loss", math.nan)),
        "train_physics_loss": float(losses.get("physics_loss", math.nan)),
        "train_total_loss": float(losses.get("total_loss", math.nan)),
        "val_mse": float(val_mse),
        "test_mse": float(test_mse),
        "test_physics_violation": float(phys_viol),
        "generalization_gap": float(test_mse - val_mse),
        "num_parameters": int(num_params),
        "model_size_bytes": int(model_size_bytes),
        "model_size_kb": float(model_size_bytes / 1024.0),
        "inference_time_ms": float(inference_ms),
        "train_seconds_until_checkpoint": float(train_seconds_so_far),
        "mse_x_params": float(test_mse * max(num_params, 1)),
        "mse_x_inference_ms": float(test_mse * max(inference_ms, 1e-12)),
    }
    return row


def train_model_with_checkpoints(
    *,
    mode: str,
    spec: ModelSpec,
    combo: SweepCombo,
    seed: int,
    data: PreparedData,
    checkpoints: list[int],
    learning_rate: float,
    dropout: float,
    device: str,
    completed_keys: set[tuple[Any, ...]],
    inference_repeats: int,
) -> list[dict[str, Any]]:
    needed_checkpoints = []
    for ep in checkpoints:
        key = make_expected_key(
            mode=mode,
            seed=seed,
            checkpoint_epoch=ep,
            combo=combo,
            spec=spec,
            learning_rate=learning_rate,
            dropout=dropout if spec.name == "FULL_GNN" else 0.0,
        )
        if key not in completed_keys:
            needed_checkpoints.append(ep)

    if not needed_checkpoints:
        return []

    max_epochs = max(checkpoints)
    set_seed(seed)
    torch.manual_seed(seed + MODEL_SEED_OFFSETS[spec.name])
    if device.startswith("cuda"):
        torch.cuda.manual_seed_all(seed + MODEL_SEED_OFFSETS[spec.name])

    model = build_model(spec, hidden_dim=combo.hidden_dim, dropout=dropout).to(device)
    optimizer = torch.optim.Adam(model.parameters(), lr=learning_rate)
    loss_fn = nn.MSELoss()

    rows: list[dict[str, Any]] = []
    train_start = time.perf_counter()
    last_losses = {"data_loss": math.nan, "physics_loss": math.nan, "total_loss": math.nan}

    for epoch in range(1, max_epochs + 1):
        last_losses = train_one_epoch(
            model=model,
            x_input=data.x_train_in,
            x_target=data.x_train_tgt,
            edge_index=data.edge_index,
            optimizer=optimizer,
            loss_fn=loss_fn,
            physics_lambda=spec.physics_lambda if spec.is_pinn else 0.0,
            grid_size=GRID_SIZE,
            alpha=ALPHA,
            dt=DT,
            dx=DX,
        )

        if epoch in checkpoints:
            train_seconds_so_far = time.perf_counter() - train_start
            row = evaluate_checkpoint(
                mode=mode,
                model=model,
                spec=spec,
                data=data,
                seed=seed,
                checkpoint_epoch=epoch,
                max_epochs=max_epochs,
                combo=combo,
                learning_rate=learning_rate,
                dropout=dropout if spec.name == "FULL_GNN" else 0.0,
                device=device,
                losses=last_losses,
                train_seconds_so_far=train_seconds_so_far,
                inference_repeats=inference_repeats,
            )
            if make_row_key(row) not in completed_keys:
                rows.append(row)
                completed_keys.add(make_row_key(row))

    del model
    if device.startswith("cuda"):
        torch.cuda.empty_cache()
    return rows


def sort_results(df: pd.DataFrame) -> pd.DataFrame:
    cols = [
        "seed",
        "checkpoint_epoch",
        "hidden_dim",
        "data_fraction",
        "noise_level",
        "model_order",
        "physics_lambda",
    ]
    existing = [c for c in cols if c in df.columns]
    ascending = [True, True, False, False, True, True, True][: len(existing)]
    return df.sort_values(existing, ascending=ascending).reset_index(drop=True)


def aggregate_mean_std(df: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    return (
        df.groupby(group_cols, dropna=False)
        .agg(
            runs=("test_mse", "count"),
            mean_test_mse=("test_mse", "mean"),
            std_test_mse=("test_mse", "std"),
            min_test_mse=("test_mse", "min"),
            mean_val_mse=("val_mse", "mean"),
            mean_physics_violation=("test_physics_violation", "mean"),
            mean_inference_ms=("inference_time_ms", "mean"),
            mean_train_seconds=("train_seconds_until_checkpoint", "mean"),
            mean_parameters=("num_parameters", "mean"),
        )
        .reset_index()
        .sort_values(group_cols)
    )


def build_tiny_vs_full(df: pd.DataFrame) -> pd.DataFrame:
    group_cols = [
        "mode",
        "seed",
        "checkpoint_epoch",
        "hidden_dim",
        "data_fraction",
        "noise_level",
        "learning_rate",
    ]
    if df.empty or "FULL_GNN" not in set(df["model"]) or "TINY_GNN" not in set(df["model"]):
        return pd.DataFrame()

    full = df[df["model"] == "FULL_GNN"][
        group_cols + ["test_mse", "test_physics_violation", "num_parameters", "inference_time_ms"]
    ].rename(
        columns={
            "test_mse": "full_test_mse",
            "test_physics_violation": "full_physics_violation",
            "num_parameters": "full_params",
            "inference_time_ms": "full_inference_ms",
        }
    )

    tiny = df[df["model"] == "TINY_GNN"][
        group_cols + ["test_mse", "test_physics_violation", "num_parameters", "inference_time_ms"]
    ].rename(
        columns={
            "test_mse": "tiny_test_mse",
            "test_physics_violation": "tiny_physics_violation",
            "num_parameters": "tiny_params",
            "inference_time_ms": "tiny_inference_ms",
        }
    )

    pinn = df[df["model"] == "TINY_GNN_PINN"].copy()
    if not pinn.empty:
        idx = pinn.groupby(group_cols)["test_mse"].idxmin()
        pinn_best = pinn.loc[idx, group_cols + ["physics_lambda", "test_mse", "test_physics_violation", "num_parameters", "inference_time_ms"]].rename(
            columns={
                "physics_lambda": "best_pinn_lambda",
                "test_mse": "best_pinn_test_mse",
                "test_physics_violation": "best_pinn_physics_violation",
                "num_parameters": "best_pinn_params",
                "inference_time_ms": "best_pinn_inference_ms",
            }
        )
    else:
        pinn_best = pd.DataFrame(columns=group_cols)

    out = full.merge(tiny, on=group_cols, how="inner").merge(pinn_best, on=group_cols, how="left")
    if out.empty:
        return out

    out["tiny_vs_full_mse_ratio"] = out["tiny_test_mse"] / out["full_test_mse"]
    out["best_pinn_vs_full_mse_ratio"] = out["best_pinn_test_mse"] / out["full_test_mse"]
    out["tiny_param_reduction_pct"] = 100.0 * (1.0 - out["tiny_params"] / out["full_params"])
    out["tiny_faster_pct"] = 100.0 * (1.0 - out["tiny_inference_ms"] / out["full_inference_ms"])

    def winner(row: pd.Series) -> str:
        values = {
            "FULL_GNN": row.get("full_test_mse", math.inf),
            "TINY_GNN": row.get("tiny_test_mse", math.inf),
            "TINY_GNN_PINN": row.get("best_pinn_test_mse", math.inf),
        }
        return min(values, key=values.get)

    out["winner_by_test_mse"] = out.apply(winner, axis=1)
    return out.sort_values(group_cols).reset_index(drop=True)


def build_report_tables(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    df = sort_results(df)

    best_runs = df.sort_values(["test_mse", "test_physics_violation", "num_parameters"]).head(200)
    summary_model_epoch = aggregate_mean_std(df, ["model", "model_display", "checkpoint_epoch"])
    summary_by_config = aggregate_mean_std(
        df,
        [
            "model",
            "model_display",
            "checkpoint_epoch",
            "hidden_dim",
            "data_fraction",
            "noise_level",
            "physics_lambda",
        ],
    )
    summary_hidden = aggregate_mean_std(df, ["model", "model_display", "hidden_dim"])
    summary_data_fraction = aggregate_mean_std(df, ["model", "model_display", "data_fraction"])
    summary_noise = aggregate_mean_std(df, ["model", "model_display", "noise_level"])
    summary_pinn_lambda = aggregate_mean_std(
        df[df["model"] == "TINY_GNN_PINN"],
        ["physics_lambda", "checkpoint_epoch", "hidden_dim", "data_fraction", "noise_level"],
    )
    tiny_vs_full = build_tiny_vs_full(df)

    # Small chart-friendly pivots.
    epoch_pivot = (
        summary_model_epoch.pivot_table(
            index="checkpoint_epoch",
            columns="model_display",
            values="mean_test_mse",
            aggfunc="mean",
        )
        .reset_index()
        .sort_values("checkpoint_epoch")
    )
    hidden_pivot = (
        summary_hidden.pivot_table(
            index="hidden_dim",
            columns="model_display",
            values="mean_test_mse",
            aggfunc="mean",
        )
        .reset_index()
        .sort_values("hidden_dim", ascending=False)
    )
    noise_pivot = (
        summary_noise.pivot_table(
            index="noise_level",
            columns="model_display",
            values="mean_test_mse",
            aggfunc="mean",
        )
        .reset_index()
        .sort_values("noise_level")
    )

    return {
        "All_Runs": df,
        "Best_Runs": best_runs,
        "Tiny_vs_Full": tiny_vs_full,
        "Summary_Model_Epoch": summary_model_epoch,
        "Summary_By_Config": summary_by_config,
        "Summary_Hidden": summary_hidden,
        "Summary_DataFraction": summary_data_fraction,
        "Summary_Noise": summary_noise,
        "Summary_PINN_Lambda": summary_pinn_lambda,
        "Chart_Epoch_Data": epoch_pivot,
        "Chart_Hidden_Data": hidden_pivot,
        "Chart_Noise_Data": noise_pivot,
    }


def safe_sheet_name(name: str) -> str:
    return name[:31]


def write_excel_report(
    *,
    csv_path: Path,
    xlsx_path: Path,
    args: argparse.Namespace,
    started_at: float,
) -> None:
    from openpyxl import load_workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    if not csv_path.exists():
        print(f"No CSV found at {csv_path}; skipping Excel report.")
        return

    df = pd.read_csv(csv_path)
    if df.empty:
        print("CSV exists but is empty; skipping Excel report.")
        return

    # Drop duplicate rows caused by retries/resume and keep the newest result.
    df = df.drop_duplicates(KEY_COLS, keep="last")
    df = sort_results(df)
    df.to_csv(csv_path, index=False)

    tables = build_report_tables(df)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        readme = pd.DataFrame(
            [
                ["Project", "TFG TinyGNN vs FullGNN heat diffusion"],
                ["Generated at", datetime.now().isoformat(timespec="seconds")],
                ["Mode", args.mode],
                ["Rows in All_Runs", len(df)],
                ["Unique seeds", df["seed"].nunique()],
                ["Epoch checkpoints", ", ".join(map(str, sorted(df["checkpoint_epoch"].unique())))],
                ["Hidden dimensions", ", ".join(map(str, sorted(df["hidden_dim"].unique(), reverse=True)))],
                ["Data fractions", ", ".join(map(str, sorted(df["data_fraction"].unique(), reverse=True)))],
                ["Noise levels", ", ".join(map(str, sorted(df["noise_level"].unique())))],
                ["PINN lambdas", ", ".join(map(str, sorted(df.loc[df["is_pinn"] == True, "physics_lambda"].unique())))],
                ["Device", args.device_resolved],
                ["Learning rate", args.learning_rate],
                ["FullGNN dropout", args.dropout],
                ["Elapsed wall time seconds", round(time.perf_counter() - started_at, 2)],
                ["Important", "Full mode is a large factorial sweep. Recommended mode is easier to justify in a TFG."],
            ],
            columns=["Field", "Value"],
        )
        readme.to_excel(writer, sheet_name="README", index=False)

        for sheet_name, table_df in tables.items():
            table_df.to_excel(writer, sheet_name=safe_sheet_name(sheet_name), index=False)

    wb = load_workbook(xlsx_path)

    # Dashboard sheet ---------------------------------------------------------
    if "Dashboard" in wb.sheetnames:
        del wb["Dashboard"]
    ws = wb.create_sheet("Dashboard", 0)

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    good_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "TFG — TinyGNN vs FullGNN experiment sweep"
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws.merge_cells("A1:H1")

    best = df.sort_values("test_mse").iloc[0]
    model_counts = df.groupby("model_display")["test_mse"].count().reset_index(name="rows")
    tiny_vs_full = tables["Tiny_vs_Full"]
    if not tiny_vs_full.empty:
        tiny_win_rate = 100.0 * (tiny_vs_full["winner_by_test_mse"] == "TINY_GNN").mean()
        pinn_win_rate = 100.0 * (tiny_vs_full["winner_by_test_mse"] == "TINY_GNN_PINN").mean()
    else:
        tiny_win_rate = math.nan
        pinn_win_rate = math.nan

    kpis = [
        ("Total result rows", len(df)),
        ("Best model", best["model_display"]),
        ("Best test MSE", best["test_mse"]),
        ("Best config", f"seed={best['seed']}, epoch={best['checkpoint_epoch']}, h={best['hidden_dim']}, data={best['data_fraction']}, noise={best['noise_level']}, λ={best['physics_lambda']}"),
        ("TinyGNN win rate vs Full", tiny_win_rate),
        ("TinyGNN+PINN win rate vs Full", pinn_win_rate),
    ]
    ws["A3"] = "KPIs"
    ws["A3"].font = Font(bold=True, size=13)
    for idx, (label, value) in enumerate(kpis, start=4):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, value)
        ws.cell(idx, 1).fill = header_fill
        ws.cell(idx, 1).font = Font(bold=True)
        ws.cell(idx, 1).border = border
        ws.cell(idx, 2).border = border
        if isinstance(value, float):
            ws.cell(idx, 2).number_format = "0.000000"

    ws["D3"] = "Rows by model"
    ws["D3"].font = Font(bold=True, size=13)
    ws.cell(4, 4, "Model")
    ws.cell(4, 5, "Rows")
    for cell in ws[4][3:5]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = border
    for i, row in enumerate(model_counts.itertuples(index=False), start=5):
        ws.cell(i, 4, row.model_display)
        ws.cell(i, 5, int(row.rows))
        ws.cell(i, 4).border = border
        ws.cell(i, 5).border = border

    ws["A12"] = "How to read this workbook"
    ws["A12"].font = Font(bold=True, size=13)
    notes = [
        "All_Runs: every evaluated checkpoint in the order seed → epoch → model.",
        "Best_Runs: globally best configurations by test MSE.",
        "Tiny_vs_Full: direct comparison for the same seed/epoch/hidden/data/noise.",
        "Summary_*: averages over repeated seeds/configurations for easier conclusions.",
        "For the TFG conclusion, focus on test_mse, physics_violation, parameters and inference_time_ms.",
    ]
    for i, note in enumerate(notes, start=13):
        ws.cell(i, 1, f"• {note}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

    # Add charts from helper sheets ------------------------------------------
    def add_line_chart(data_sheet_name: str, anchor: str, title: str, x_title: str, y_title: str) -> None:
        if data_sheet_name not in wb.sheetnames:
            return
        ds = wb[data_sheet_name]
        if ds.max_row < 2 or ds.max_column < 2:
            return
        chart = LineChart()
        chart.title = title
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title
        chart.height = 8
        chart.width = 16
        data_ref = Reference(ds, min_col=2, max_col=ds.max_column, min_row=1, max_row=ds.max_row)
        cats_ref = Reference(ds, min_col=1, min_row=2, max_row=ds.max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.style = 13
        ws.add_chart(chart, anchor)

    add_line_chart("Chart_Epoch_Data", "A20", "Mean test MSE by epoch", "Epoch", "Mean test MSE")
    add_line_chart("Chart_Hidden_Data", "J3", "Mean test MSE by hidden dimension", "Hidden dim", "Mean test MSE")
    add_line_chart("Chart_Noise_Data", "J20", "Mean test MSE by noise level", "Noise", "Mean test MSE")

    # General formatting ------------------------------------------------------
    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False

        # Header styling and autofilter/table.
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = title_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            if sheet.title != "Dashboard" and sheet.max_row >= 2:
                ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
                table_name = "tbl_" + "".join(ch for ch in sheet.title if ch.isalnum())[:20]
                try:
                    tab = Table(displayName=table_name, ref=ref)
                    style = TableStyleInfo(
                        name="TableStyleMedium2",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False,
                    )
                    tab.tableStyleInfo = style
                    sheet.add_table(tab)
                except Exception:
                    sheet.auto_filter.ref = ref

        # Column widths.
        for col_idx in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, min(sheet.max_row, 1000) + 1):
                value = sheet.cell(row_idx, col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            sheet.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 38)

        # Numeric formats.
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                header = sheet.cell(1, cell.column).value
                if header is None:
                    continue
                header_s = str(header).lower()
                if isinstance(cell.value, float):
                    if "mse" in header_s or "violation" in header_s or "loss" in header_s or "gap" in header_s:
                        cell.number_format = "0.000000E+00"
                    elif "time" in header_s or "seconds" in header_s or "ms" in header_s:
                        cell.number_format = "0.0000"
                    elif "pct" in header_s or "rate" in header_s:
                        cell.number_format = "0.00"
                    else:
                        cell.number_format = "0.000000"

        # Conditional formatting for important numeric columns.
        if sheet.title != "Dashboard" and sheet.max_row >= 3:
            headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
            for important in ["test_mse", "val_mse", "test_physics_violation", "mean_test_mse", "full_test_mse", "tiny_test_mse", "best_pinn_test_mse"]:
                if important in headers:
                    c_idx = headers.index(important) + 1
                    col = get_column_letter(c_idx)
                    rng = f"{col}2:{col}{sheet.max_row}"
                    sheet.conditional_formatting.add(
                        rng,
                        ColorScaleRule(
                            start_type="min",
                            start_color="63BE7B",
                            mid_type="percentile",
                            mid_value=50,
                            mid_color="FFEB84",
                            end_type="max",
                            end_color="F8696B",
                        ),
                    )

    # Dashboard specific sizing.
    ws.freeze_panes = None
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 12
    ws.row_dimensions[1].height = 26

    wb.save(xlsx_path)


def estimate_workload(
    *,
    combos: list[SweepCombo],
    seeds: list[int],
    model_specs: list[ModelSpec],
    checkpoints: list[int],
) -> dict[str, int]:
    trainings = len(combos) * len(seeds) * len(model_specs)
    result_rows = trainings * len(checkpoints)
    max_epochs = max(checkpoints)
    total_epoch_loops = trainings * max_epochs
    return {
        "combos": len(combos),
        "seeds": len(seeds),
        "model_trainings": trainings,
        "result_rows": result_rows,
        "max_epochs_per_training": max_epochs,
        "total_epoch_loops": total_epoch_loops,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run TFG TinyGNN/FullGNN sweep and export a styled Excel report.")
    parser.add_argument("--mode", choices=["smoke", "recommended", "full"], default="recommended")
    parser.add_argument("--epochs", default=",".join(map(str, DEFAULT_EPOCHS)), help="Comma list, e.g. 1,10,50,100,300")
    parser.add_argument("--seeds", default="40:50", help="Comma list or inclusive range, e.g. 40:50")
    parser.add_argument("--hidden-dims", default=",".join(map(str, DEFAULT_HIDDEN_DIMS)))
    parser.add_argument("--data-fractions", default=",".join(map(str, DEFAULT_DATA_FRACTIONS)))
    parser.add_argument("--noise-levels", default=",".join(map(str, DEFAULT_NOISE_LEVELS)))
    parser.add_argument("--physics-lambdas", default=",".join(map(str, DEFAULT_PHYSICS_LAMBDAS)))
    parser.add_argument("--learning-rate", type=float, default=1e-3)
    parser.add_argument("--dropout", type=float, default=0.1, help="Only used by FullGNN")
    parser.add_argument("--device", default="auto", help="auto, cuda, cuda:0, cpu")
    parser.add_argument("--output-dir", default=str(CONFIG.sweep_results_dir))
    parser.add_argument("--only-models", default="", help="Optional comma list: MLP,FULL_GNN,TINY_GNN,TINY_GNN_PINN")
    parser.add_argument("--inference-repeats", type=int, default=20)
    parser.add_argument("--limit-trainings", type=int, default=0, help="Debug: stop after N model trainings. 0 = no limit.")
    parser.add_argument("--no-excel", action="store_true", help="Only write/update CSV; skip styled Excel generation.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    started_at = time.perf_counter()

    checkpoints = sorted(unique_preserve_order(parse_int_list(args.epochs)))
    seeds = parse_int_list(args.seeds)
    hidden_dims = parse_int_list(args.hidden_dims)
    data_fractions = parse_float_list(args.data_fractions)
    noise_levels = parse_float_list(args.noise_levels)
    physics_lambdas = parse_float_list(args.physics_lambdas)
    only_models = [x.strip().upper() for x in args.only_models.split(",") if x.strip()]

    if args.mode == "smoke":
        checkpoints = [1]
        seeds = [40]
        hidden_dims = [16]
        data_fractions = [1.0]
        noise_levels = [0.0]
        physics_lambdas = [0.1]

    device = resolve_device(args.device)
    args.device_resolved = device
    configure_torch(device)

    combos = make_sweep_combos(args.mode, hidden_dims, data_fractions, noise_levels)
    model_specs = make_model_specs(only_models, physics_lambdas)
    workload = estimate_workload(combos=combos, seeds=seeds, model_specs=model_specs, checkpoints=checkpoints)

    output_dir = Path(args.output_dir)
    csv_path = output_dir / "all_results.csv"
    xlsx_path = output_dir / "sweep_report.xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)

    print("\n" + "=" * 88)
    print("TFG TinyGNN vs FullGNN sweep")
    print("=" * 88)
    print(f"Mode:                 {args.mode}")
    print(f"Device:               {device}")
    print(f"Epoch checkpoints:    {checkpoints}")
    print(f"Seeds:                {seeds}")
    print(f"Hidden dims:          {hidden_dims}")
    print(f"Data fractions:       {data_fractions}")
    print(f"Noise levels:         {noise_levels}")
    print(f"PINN lambdas:         {physics_lambdas}")
    print(f"Output CSV:           {csv_path}")
    print(f"Output Excel:         {xlsx_path}")
    print("-" * 88)
    print(f"Unique h/data/noise combos: {workload['combos']}")
    print(f"Model trainings:            {workload['model_trainings']}")
    print(f"Rows in Excel/CSV:           {workload['result_rows']}")
    print(f"Total epoch loops:           {workload['total_epoch_loops']}")
    print("=" * 88 + "\n")

    if args.mode == "full":
        print(
            "WARNING: --mode full is the exact factorial sweep, but it is very large. "
            "If it is too slow, stop it and run --mode recommended; the CSV is resumable.\n"
        )

    completed_keys = load_completed_keys(csv_path)
    if completed_keys:
        print(f"Resume enabled: found {len(completed_keys)} completed result rows in {csv_path}\n")

    training_counter = 0
    for seed in seeds:
        for combo_idx, combo in enumerate(combos, start=1):
            print(
                f"[seed {seed}] combo {combo_idx}/{len(combos)} | "
                f"hidden={combo.hidden_dim}, data_fraction={combo.data_fraction}, noise={combo.noise_level}"
            )
            data = prepare_data(
                seed=seed,
                data_fraction=combo.data_fraction,
                noise_level=combo.noise_level,
                device=device,
            )

            for spec in model_specs:
                expected_keys = [
                    make_expected_key(
                        mode=args.mode,
                        seed=seed,
                        checkpoint_epoch=ep,
                        combo=combo,
                        spec=spec,
                        learning_rate=args.learning_rate,
                        dropout=args.dropout if spec.name == "FULL_GNN" else 0.0,
                    )
                    for ep in checkpoints
                ]
                if all(k in completed_keys for k in expected_keys):
                    print(f"  SKIP {spec.display_name:15s} lambda={spec.physics_lambda:g} (already done)")
                    continue

                training_counter += 1
                if args.limit_trainings and training_counter > args.limit_trainings:
                    print("Reached --limit-trainings; stopping early.")
                    if not args.no_excel:
                        write_excel_report(csv_path=csv_path, xlsx_path=xlsx_path, args=args, started_at=started_at)
                    return

                print(f"  TRAIN {spec.display_name:15s} lambda={spec.physics_lambda:g}")
                rows = train_model_with_checkpoints(
                    mode=args.mode,
                    spec=spec,
                    combo=combo,
                    seed=seed,
                    data=data,
                    checkpoints=checkpoints,
                    learning_rate=args.learning_rate,
                    dropout=args.dropout,
                    device=device,
                    completed_keys=completed_keys,
                    inference_repeats=args.inference_repeats,
                )
                append_rows_to_csv(rows, csv_path)
                for row in rows:
                    print(
                        f"    epoch={row['checkpoint_epoch']:>3d} | "
                        f"test_mse={row['test_mse']:.6e} | "
                        f"phys={row['test_physics_violation']:.6e} | "
                        f"params={row['num_parameters']} | "
                        f"infer={row['inference_time_ms']:.4f} ms"
                    )

            # Release data tensors between combos.
            del data
            if device.startswith("cuda"):
                torch.cuda.empty_cache()

    if not args.no_excel:
        print("\nBuilding styled Excel report...")
        write_excel_report(csv_path=csv_path, xlsx_path=xlsx_path, args=args, started_at=started_at)

    print("\nDone.")
    print(f"CSV:   {csv_path}")
    if not args.no_excel:
        print(f"Excel: {xlsx_path}")


if __name__ == "__main__":
    main()
