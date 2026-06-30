#!/usr/bin/env python3
"""
Corrected TFG experiment sweep.

This version addresses the methodological issues detected by the supervisor:

1. Controlled physics mismatch: data can be generated with data_alpha and a
   small reaction_mu term, while PINN enforces pure diffusion with pinn_alpha.
2. Real generalisation: train/validation/test are independent trajectories.
3. Random initial conditions: position, amplitude and width vary per trajectory.
4. data_fraction subsamples complete training trajectories.
5. PINN lambda sensitivity is explicit and recorded.
6. Excel report includes 95% confidence intervals and paired t-tests.
7. Inference time is mean/std over many repeated forward passes.
8. Early stopping and model selection are based on validation MSE.

Recommended quick checks:
    python src/run_sweep_experiments_v2.py --mode smoke
    python src/run_sweep_experiments_v2.py --mode recommended

A heavier run:
    python src/run_sweep_experiments_v2.py --mode full --seeds 40:50
"""

from __future__ import annotations

import argparse
import copy
import itertools
import math
import os
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

THIS_FILE = Path(__file__).resolve()
THIS_DIR = THIS_FILE.parent
if (THIS_DIR / "config.py").exists():
    sys.path.insert(0, str(THIS_DIR))
elif (THIS_DIR / "src" / "config.py").exists():
    sys.path.insert(0, str(THIS_DIR / "src"))

import pandas as pd
import torch
from torch import nn

from config import ALPHA, CONFIG, DT, DX, GRID_SIZE
from data.graph_builder import build_grid_graph
from data.heat_diffusion import (
    add_gaussian_noise,
    build_temporal_pairs,
    simulate_heat_diffusion_trajectories,
)
from models.baselines import MLPBaseline
from models.full_gnn import FullGNN
from models.tiny_gnn import TinyGNN
from training.evaluator import (
    evaluate_physics_violation,
    evaluate_prediction_error,
    measure_inference_time_stats_ms,
    measure_model_size_bytes,
)
from training.trainer import train_one_epoch
from utils.metrics import count_trainable_parameters
from utils.seed import set_seed


DEFAULT_CHECKPOINTS = [50, 100, 300]
DEFAULT_SEEDS = list(range(40, 51))
DEFAULT_HIDDEN_DIMS = [32, 16, 8, 4]
DEFAULT_DATA_FRACTIONS = [1.0, 0.5, 0.25, 0.125]
DEFAULT_NOISE_LEVELS = [0.0, 0.05, 0.10, 0.20]
DEFAULT_PHYSICS_LAMBDAS = [0.001, 0.01, 0.1, 1.0, 10.0, 30.0]

MODEL_ORDER = {"MLP": 0, "FULL_GNN": 1, "TINY_GNN": 2, "TINY_GNN_PINN": 3}
MODEL_SEED_OFFSETS = {"MLP": 101, "FULL_GNN": 202, "TINY_GNN": 303, "TINY_GNN_PINN": 303}

KEY_COLS = [
    "mode",
    "seed",
    "checkpoint_epoch",
    "selected_epoch",
    "hidden_dim",
    "data_fraction",
    "noise_level",
    "model",
    "physics_lambda",
    "learning_rate",
    "dropout",
    "data_alpha",
    "pinn_alpha",
    "reaction_mu",
    "trajectory_timesteps",
    "train_trajectories_total",
    "train_trajectories_kept",
    "val_trajectories",
    "test_trajectories",
]


@dataclass(frozen=True)
class PreparedData:
    x_train_in: torch.Tensor
    x_train_tgt: torch.Tensor
    x_val_in: torch.Tensor
    x_val_tgt: torch.Tensor
    x_test_in: torch.Tensor
    x_test_tgt: torch.Tensor
    edge_index: torch.Tensor
    train_pairs: int
    val_pairs: int
    test_pairs: int
    train_trajectories_total: int
    train_trajectories_kept: int
    val_trajectories: int
    test_trajectories: int
    data_alpha: float
    pinn_alpha: float
    reaction_mu: float
    trajectory_timesteps: int


@dataclass(frozen=True)
class SweepCombo:
    hidden_dim: int
    data_fraction: float
    noise_level: float


@dataclass(frozen=True)
class ModelSpec:
    name: str
    display_name: str
    is_pinn: bool
    physics_lambda: float


def parse_int_list(value: str) -> list[int]:
    value = value.strip()
    if ":" in value:
        start, end = value.split(":", 1)
        return list(range(int(start), int(end) + 1))
    return [int(x.strip()) for x in value.split(",") if x.strip()]


def parse_float_list(value: str) -> list[float]:
    return [float(x.strip()) for x in value.split(",") if x.strip()]


def unique_preserve_order(items: Iterable[Any]) -> list[Any]:
    seen: set[Any] = set()
    out: list[Any] = []
    for item in items:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out


def resolve_device(device_arg: str) -> str:
    if device_arg == "auto":
        return "cuda" if torch.cuda.is_available() else "cpu"
    if device_arg.startswith("cuda") and not torch.cuda.is_available():
        raise RuntimeError("You requested CUDA, but torch.cuda.is_available() is False.")
    return device_arg


def configure_torch(device: str) -> None:
    if device.startswith("cuda"):
        torch.backends.cudnn.benchmark = True
        try:
            torch.set_float32_matmul_precision("high")
        except Exception:
            pass


def make_sweep_combos(mode: str, hidden_dims: list[int], data_fractions: list[float], noise_levels: list[float]) -> list[SweepCombo]:
    if mode == "smoke":
        return [SweepCombo(hidden_dim=16, data_fraction=1.0, noise_level=0.0)]
    if mode == "full":
        return [SweepCombo(h, f, n) for h, f, n in itertools.product(hidden_dims, data_fractions, noise_levels)]
    if mode == "recommended":
        anchor_hidden = 16 if 16 in hidden_dims else hidden_dims[0]
        triples: list[tuple[int, float, float]] = []
        triples.extend((h, 1.0, 0.0) for h in hidden_dims)
        triples.extend((anchor_hidden, f, 0.0) for f in data_fractions)
        triples.extend((anchor_hidden, 1.0, n) for n in noise_levels)
        return [SweepCombo(h, f, n) for h, f, n in unique_preserve_order(triples)]
    raise ValueError(f"Unknown mode: {mode}")


def make_model_specs(only_models: list[str], physics_lambdas: list[float]) -> list[ModelSpec]:
    requested = {m.upper() for m in only_models}

    def enabled(name: str) -> bool:
        return not requested or name in requested

    specs: list[ModelSpec] = []
    if enabled("MLP"):
        specs.append(ModelSpec("MLP", "MLP", is_pinn=False, physics_lambda=0.0))
    if enabled("FULL_GNN"):
        specs.append(ModelSpec("FULL_GNN", "FULL_GNN", is_pinn=False, physics_lambda=0.0))
    if enabled("TINY_GNN"):
        specs.append(ModelSpec("TINY_GNN", "TINY_GNN", is_pinn=False, physics_lambda=0.0))
    if enabled("TINY_GNN_PINN"):
        for lam in physics_lambdas:
            specs.append(ModelSpec("TINY_GNN_PINN", "TINY_GNN + PINN", is_pinn=True, physics_lambda=float(lam)))
    if not specs:
        raise ValueError("No models selected. Use: MLP,FULL_GNN,TINY_GNN,TINY_GNN_PINN")
    return specs


def build_model(spec: ModelSpec, hidden_dim: int, dropout: float) -> nn.Module:
    if spec.name == "MLP":
        return MLPBaseline(hidden_dim=hidden_dim)
    if spec.name == "FULL_GNN":
        return FullGNN(hidden_dim=hidden_dim, dropout=dropout)
    if spec.name in {"TINY_GNN", "TINY_GNN_PINN"}:
        return TinyGNN(hidden_dim=hidden_dim)
    raise ValueError(f"Unknown model: {spec.name}")


def prepare_data(*, seed: int, combo: SweepCombo, args: argparse.Namespace, device: str) -> PreparedData:
    set_seed(seed)

    sim = simulate_heat_diffusion_trajectories(
        grid_size=args.grid_size,
        num_timesteps=args.trajectory_timesteps,
        num_train_trajectories=args.train_trajectories,
        num_val_trajectories=args.val_trajectories,
        num_test_trajectories=args.test_trajectories,
        data_fraction=combo.data_fraction,
        data_alpha=args.data_alpha,
        pinn_alpha=args.pinn_alpha,
        reaction_mu=args.reaction_mu,
        dt=args.dt,
        dx=args.dx,
        seed=seed,
    )

    x_train_in, x_train_tgt = build_temporal_pairs(sim.train_states)
    x_val_in, x_val_tgt = build_temporal_pairs(sim.val_states)
    x_test_in, x_test_tgt = build_temporal_pairs(sim.test_states)

    # Noise only affects training supervision. Validation/test stay clean.
    if combo.noise_level > 0.0:
        x_train_in = add_gaussian_noise(x_train_in, sigma=combo.noise_level, seed=seed + 17)
        x_train_tgt = add_gaussian_noise(x_train_tgt, sigma=combo.noise_level, seed=seed + 31)

    x_train_in = x_train_in.unsqueeze(-1).to(device)
    x_train_tgt = x_train_tgt.unsqueeze(-1).to(device)
    x_val_in = x_val_in.unsqueeze(-1).to(device)
    x_val_tgt = x_val_tgt.unsqueeze(-1).to(device)
    x_test_in = x_test_in.unsqueeze(-1).to(device)
    x_test_tgt = x_test_tgt.unsqueeze(-1).to(device)

    graph = build_grid_graph(grid_size=args.grid_size)
    edge_index = graph.edge_index.to(device)

    return PreparedData(
        x_train_in=x_train_in,
        x_train_tgt=x_train_tgt,
        x_val_in=x_val_in,
        x_val_tgt=x_val_tgt,
        x_test_in=x_test_in,
        x_test_tgt=x_test_tgt,
        edge_index=edge_index,
        train_pairs=int(x_train_in.shape[0]),
        val_pairs=int(x_val_in.shape[0]),
        test_pairs=int(x_test_in.shape[0]),
        train_trajectories_total=sim.num_train_trajectories_total,
        train_trajectories_kept=sim.num_train_trajectories_kept,
        val_trajectories=sim.num_val_trajectories,
        test_trajectories=sim.num_test_trajectories,
        data_alpha=sim.data_alpha,
        pinn_alpha=sim.pinn_alpha,
        reaction_mu=sim.reaction_mu,
        trajectory_timesteps=args.trajectory_timesteps,
    )


def cuda_sync_if_needed(device: str) -> None:
    if device.startswith("cuda") and torch.cuda.is_available():
        torch.cuda.synchronize()


def normalize_key_value(value: Any) -> Any:
    """Canonicalise values used for resume/dedup keys.

    CSV round-tripping turns values such as 0.08000000000000002 into 0.08.
    Exact tuple comparison then fails and the sweep retrains already completed
    configurations.  Rounding float-like values makes resume stable while still
    keeping enough precision for these hyperparameters.
    """
    if pd.isna(value):
        return None
    if isinstance(value, float):
        return round(value, 12)
    return value


def make_row_key(row: dict[str, Any]) -> tuple[Any, ...]:
    return tuple(normalize_key_value(row[col]) for col in KEY_COLS)


def make_df_key(row: pd.Series) -> tuple[Any, ...]:
    return tuple(normalize_key_value(row[col]) for col in KEY_COLS)


def load_completed_keys(csv_path: Path) -> set[tuple[Any, ...]]:
    if not csv_path.exists():
        return set()
    try:
        df = pd.read_csv(csv_path)
    except Exception:
        return set()
    if df.empty or not set(KEY_COLS).issubset(df.columns):
        return set()
    return {make_df_key(row) for _, row in df[KEY_COLS].iterrows()}


def append_rows_to_csv(rows: list[dict[str, Any]], csv_path: Path) -> None:
    if not rows:
        return
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows)
    write_header = not csv_path.exists()
    df.to_csv(csv_path, mode="a", header=write_header, index=False)


def expected_already_done(
    *,
    completed: set[tuple[Any, ...]],
    mode: str,
    seed: int,
    checkpoints: list[int],
    combo: SweepCombo,
    spec: ModelSpec,
    args: argparse.Namespace,
    selected_epoch_unknown: bool = True,
) -> bool:
    # With early stopping the selected_epoch is only known after training.
    # Therefore skip based on a softer subset to keep resume practical.
    if not completed:
        return False
    # A conservative check using all non-selected fields.
    completed_df_keys = [k for k in completed]
    needed_without_selected = []
    for ep in checkpoints:
        prefix = {
            "mode": mode,
            "seed": seed,
            "checkpoint_epoch": ep,
            "hidden_dim": combo.hidden_dim,
            "data_fraction": combo.data_fraction,
            "noise_level": combo.noise_level,
            "model": spec.name,
            "physics_lambda": spec.physics_lambda,
            "learning_rate": args.learning_rate,
            "dropout": args.dropout if spec.name == "FULL_GNN" else 0.0,
            "data_alpha": args.data_alpha,
            "pinn_alpha": args.pinn_alpha,
            "reaction_mu": args.reaction_mu,
            "trajectory_timesteps": args.trajectory_timesteps,
            "train_trajectories_total": args.train_trajectories,
            "val_trajectories": args.val_trajectories,
            "test_trajectories": args.test_trajectories,
        }
        needed_without_selected.append(prefix)

    def row_matches_key(prefix: dict[str, Any], key: tuple[Any, ...]) -> bool:
        d = dict(zip(KEY_COLS, key))
        for name, value in prefix.items():
            if d.get(name) != normalize_key_value(value):
                return False
        # train_trajectories_kept depends on data_fraction.
        expected_kept = max(1, int(round(args.train_trajectories * combo.data_fraction)))
        return d.get("train_trajectories_kept") == expected_kept

    return all(any(row_matches_key(prefix, key) for key in completed_df_keys) for prefix in needed_without_selected)


def evaluate_selected_model(
    *,
    mode: str,
    model: nn.Module,
    spec: ModelSpec,
    data: PreparedData,
    seed: int,
    checkpoint_epoch: int,
    selected_epoch: int,
    trained_epochs: int,
    max_epochs: int,
    combo: SweepCombo,
    args: argparse.Namespace,
    device: str,
    losses: dict[str, float],
    best_val_mse: float,
    train_seconds_so_far: float,
    stopped_early: bool,
) -> dict[str, Any]:
    cuda_sync_if_needed(device)

    val_mse = evaluate_prediction_error(model, data.x_val_in, data.x_val_tgt, data.edge_index)
    test_mse = evaluate_prediction_error(model, data.x_test_in, data.x_test_tgt, data.edge_index)
    phys_viol = evaluate_physics_violation(
        model,
        data.x_test_in,
        data.edge_index,
        alpha=data.pinn_alpha,
        dt=args.dt,
        dx=args.dx,
        grid_size=args.grid_size,
    )
    infer = measure_inference_time_stats_ms(
        model,
        data.x_test_in,
        data.edge_index,
        repeats=args.inference_repeats,
        warmup=args.inference_warmup,
    )
    num_params = count_trainable_parameters(model)
    model_size_bytes = measure_model_size_bytes(model)

    row = {
        "run_label": (
            f"{spec.display_name} | seed {seed} | h {combo.hidden_dim} | "
            f"data {combo.data_fraction} | noise {combo.noise_level} | λ {spec.physics_lambda} | "
            f"selected epoch {selected_epoch}"
        ),
        "mode": mode,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "model": spec.name,
        "model_display": spec.display_name,
        "model_order": MODEL_ORDER[spec.name],
        "is_pinn": bool(spec.is_pinn),
        "seed": int(seed),
        "checkpoint_epoch": int(checkpoint_epoch),
        "selected_epoch": int(selected_epoch),
        "trained_epochs": int(trained_epochs),
        "max_epochs": int(max_epochs),
        "stopped_early": bool(stopped_early),
        "hidden_dim": int(combo.hidden_dim),
        "data_fraction": float(combo.data_fraction),
        "noise_level": float(combo.noise_level),
        "physics_lambda": float(spec.physics_lambda),
        "learning_rate": float(args.learning_rate),
        "dropout": float(args.dropout if spec.name == "FULL_GNN" else 0.0),
        "device": device,
        "grid_size": int(args.grid_size),
        "trajectory_timesteps": int(data.trajectory_timesteps),
        "train_trajectories_total": int(data.train_trajectories_total),
        "train_trajectories_kept": int(data.train_trajectories_kept),
        "val_trajectories": int(data.val_trajectories),
        "test_trajectories": int(data.test_trajectories),
        "train_pairs": int(data.train_pairs),
        "val_pairs": int(data.val_pairs),
        "test_pairs": int(data.test_pairs),
        "data_alpha": float(data.data_alpha),
        "pinn_alpha": float(data.pinn_alpha),
        "reaction_mu": float(data.reaction_mu),
        "dt": float(args.dt),
        "dx": float(args.dx),
        "train_data_loss": float(losses.get("data_loss", math.nan)),
        "train_physics_loss": float(losses.get("physics_loss", math.nan)),
        "train_total_loss": float(losses.get("total_loss", math.nan)),
        "best_val_mse": float(best_val_mse),
        "val_mse": float(val_mse),
        "test_mse": float(test_mse),
        "test_physics_violation": float(phys_viol),
        "generalization_gap": float(test_mse - val_mse),
        "num_parameters": int(num_params),
        "model_size_bytes": int(model_size_bytes),
        "model_size_kb": float(model_size_bytes / 1024.0),
        "inference_time_ms": float(infer["mean_ms"]),
        "inference_time_std_ms": float(infer["std_ms"]),
        "inference_time_min_ms": float(infer["min_ms"]),
        "inference_time_max_ms": float(infer["max_ms"]),
        "inference_repeats": int(infer["repeats"]),
        "train_seconds_until_checkpoint": float(train_seconds_so_far),
        "mse_x_params": float(test_mse * max(num_params, 1)),
        "mse_x_inference_ms": float(test_mse * max(infer["mean_ms"], 1e-12)),
    }
    return row


def train_model_with_selection(
    *,
    mode: str,
    spec: ModelSpec,
    combo: SweepCombo,
    seed: int,
    data: PreparedData,
    checkpoints: list[int],
    args: argparse.Namespace,
    device: str,
) -> list[dict[str, Any]]:
    max_epochs = max(checkpoints)

    set_seed(seed)
    torch.manual_seed(seed + MODEL_SEED_OFFSETS[spec.name])
    if device.startswith("cuda"):
        torch.cuda.manual_seed_all(seed + MODEL_SEED_OFFSETS[spec.name])

    model = build_model(spec, hidden_dim=combo.hidden_dim, dropout=args.dropout).to(device)
    optimizer = torch.optim.Adam(model.parameters(), lr=args.learning_rate)
    loss_fn = nn.MSELoss()

    best_val_mse = math.inf
    best_epoch = 0
    best_state: dict[str, torch.Tensor] | None = None
    epochs_without_improvement = 0
    stopped_early = False
    last_losses: dict[str, float] = {"data_loss": math.nan, "physics_loss": math.nan, "total_loss": math.nan}
    rows: list[dict[str, Any]] = []
    train_start = time.perf_counter()

    for epoch in range(1, max_epochs + 1):
        last_losses = train_one_epoch(
            model=model,
            x_input=data.x_train_in,
            x_target=data.x_train_tgt,
            edge_index=data.edge_index,
            optimizer=optimizer,
            loss_fn=loss_fn,
            physics_lambda=spec.physics_lambda if spec.is_pinn else 0.0,
            grid_size=args.grid_size,
            alpha=data.pinn_alpha,
            dt=args.dt,
            dx=args.dx,
        )

        current_val_mse = evaluate_prediction_error(model, data.x_val_in, data.x_val_tgt, data.edge_index)
        if current_val_mse < best_val_mse - args.min_delta:
            best_val_mse = current_val_mse
            best_epoch = epoch
            best_state = copy.deepcopy(model.state_dict())
            epochs_without_improvement = 0
        else:
            epochs_without_improvement += 1

        if epoch in checkpoints:
            # Evaluate the best validation model seen up to this checkpoint, not
            # blindly the last epoch model.
            current_state = copy.deepcopy(model.state_dict())
            if best_state is not None:
                model.load_state_dict(best_state)
            rows.append(evaluate_selected_model(
                mode=mode,
                model=model,
                spec=spec,
                data=data,
                seed=seed,
                checkpoint_epoch=epoch,
                selected_epoch=best_epoch,
                trained_epochs=epoch,
                max_epochs=max_epochs,
                combo=combo,
                args=args,
                device=device,
                losses=last_losses,
                best_val_mse=best_val_mse,
                train_seconds_so_far=time.perf_counter() - train_start,
                stopped_early=False,
            ))
            model.load_state_dict(current_state)

        if args.early_stopping and epoch >= args.min_epochs and epochs_without_improvement >= args.patience:
            stopped_early = True
            break

    trained_epochs = epoch
    if stopped_early:
        # Fill missing future checkpoint rows with the selected best model so the
        # CSV remains complete and resumable.
        if best_state is not None:
            model.load_state_dict(best_state)
        already_written = {r["checkpoint_epoch"] for r in rows}
        for checkpoint in checkpoints:
            if checkpoint not in already_written:
                rows.append(evaluate_selected_model(
                    mode=mode,
                    model=model,
                    spec=spec,
                    data=data,
                    seed=seed,
                    checkpoint_epoch=checkpoint,
                    selected_epoch=best_epoch,
                    trained_epochs=trained_epochs,
                    max_epochs=max_epochs,
                    combo=combo,
                    args=args,
                    device=device,
                    losses=last_losses,
                    best_val_mse=best_val_mse,
                    train_seconds_so_far=time.perf_counter() - train_start,
                    stopped_early=True,
                ))

    del model
    if device.startswith("cuda"):
        torch.cuda.empty_cache()
    return sorted(rows, key=lambda r: r["checkpoint_epoch"])


def sort_results(df: pd.DataFrame) -> pd.DataFrame:
    cols = ["seed", "checkpoint_epoch", "hidden_dim", "data_fraction", "noise_level", "model_order", "physics_lambda"]
    existing = [c for c in cols if c in df.columns]
    ascending = [True, True, False, False, True, True, True][: len(existing)]
    return df.sort_values(existing, ascending=ascending).reset_index(drop=True)


def aggregate_with_ci(df: pd.DataFrame, group_cols: list[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    g = df.groupby(group_cols, dropna=False)
    out = g.agg(
        runs=("test_mse", "count"),
        mean_test_mse=("test_mse", "mean"),
        std_test_mse=("test_mse", "std"),
        median_test_mse=("test_mse", "median"),
        min_test_mse=("test_mse", "min"),
        p90_test_mse=("test_mse", lambda s: s.quantile(0.90)),
        mean_val_mse=("val_mse", "mean"),
        mean_physics_violation=("test_physics_violation", "mean"),
        mean_inference_ms=("inference_time_ms", "mean"),
        std_inference_ms=("inference_time_ms", "std"),
        mean_train_seconds=("train_seconds_until_checkpoint", "mean"),
        mean_parameters=("num_parameters", "mean"),
        mean_selected_epoch=("selected_epoch", "mean"),
    ).reset_index()
    out["sem_test_mse"] = out["std_test_mse"] / out["runs"].pow(0.5)
    out["ci95_low_test_mse"] = out["mean_test_mse"] - 1.96 * out["sem_test_mse"]
    out["ci95_high_test_mse"] = out["mean_test_mse"] + 1.96 * out["sem_test_mse"]
    return out.sort_values(group_cols).reset_index(drop=True)


def build_paired_ttests(df: pd.DataFrame) -> pd.DataFrame:
    try:
        from scipy import stats
    except Exception:
        stats = None

    final_epoch = int(df["checkpoint_epoch"].max()) if not df.empty else 0
    final = df[df["checkpoint_epoch"] == final_epoch].copy()
    if final.empty:
        return pd.DataFrame()

    # Compare each model to FULL_GNN and TINY_GNN, paired by identical
    # experimental setting and seed. For PINN, each lambda remains separate.
    base_cols = [
        "mode", "seed", "hidden_dim", "data_fraction", "noise_level",
        "learning_rate", "data_alpha", "pinn_alpha", "reaction_mu",
        "trajectory_timesteps", "train_trajectories_total", "train_trajectories_kept",
        "val_trajectories", "test_trajectories",
    ]

    rows: list[dict[str, Any]] = []
    candidates = final[[*base_cols, "model", "model_display", "physics_lambda", "test_mse"]]
    for baseline in ["MLP", "FULL_GNN", "TINY_GNN"]:
        left = candidates[candidates["model"] == baseline].rename(columns={
            "test_mse": "baseline_test_mse",
            "model_display": "baseline_display",
            "physics_lambda": "baseline_lambda",
        })
        if left.empty:
            continue
        for model_name in sorted(candidates["model"].unique()):
            if model_name == baseline:
                continue
            right = candidates[candidates["model"] == model_name].rename(columns={
                "test_mse": "candidate_test_mse",
                "model_display": "candidate_display",
                "physics_lambda": "candidate_lambda",
            })
            merged = left.merge(right, on=base_cols, how="inner", suffixes=("", "_candidate"))
            if merged.empty:
                continue
            # Difference < 0 means candidate has lower MSE than baseline.
            diff = merged["candidate_test_mse"] - merged["baseline_test_mse"]
            n = len(diff)
            mean_diff = float(diff.mean())
            std_diff = float(diff.std(ddof=1)) if n > 1 else math.nan
            sem = std_diff / math.sqrt(n) if n > 1 and not math.isnan(std_diff) else math.nan
            ci_low = mean_diff - 1.96 * sem if not math.isnan(sem) else math.nan
            ci_high = mean_diff + 1.96 * sem if not math.isnan(sem) else math.nan
            if stats is not None and n > 1:
                t_stat, p_value = stats.ttest_rel(merged["candidate_test_mse"], merged["baseline_test_mse"])
                t_stat = float(t_stat)
                p_value = float(p_value)
            else:
                t_stat, p_value = math.nan, math.nan
            rows.append({
                "checkpoint_epoch": final_epoch,
                "baseline_model": baseline,
                "candidate_model": model_name,
                "candidate_lambda_values": ",".join(map(str, sorted(merged["candidate_lambda"].unique()))),
                "paired_runs": int(n),
                "mean_candidate_minus_baseline": mean_diff,
                "std_diff": std_diff,
                "ci95_low_diff": ci_low,
                "ci95_high_diff": ci_high,
                "t_statistic": t_stat,
                "p_value": p_value,
                "candidate_better_rate": float((diff < 0).mean()),
            })
    return pd.DataFrame(rows)


def build_lambda_selection_table(df: pd.DataFrame) -> pd.DataFrame:
    pinn = df[(df["model"] == "TINY_GNN_PINN") & (df["checkpoint_epoch"] == df["checkpoint_epoch"].max())].copy()
    if pinn.empty:
        return pd.DataFrame()

    group_cols = ["seed", "hidden_dim", "data_fraction", "noise_level", "data_alpha", "pinn_alpha", "reaction_mu"]
    idx = pinn.groupby(group_cols)["val_mse"].idxmin()
    selected = pinn.loc[idx].copy()
    return (
        selected.groupby(["hidden_dim", "data_fraction", "noise_level", "physics_lambda"])
        .agg(
            selections=("test_mse", "count"),
            mean_val_mse=("val_mse", "mean"),
            mean_test_mse=("test_mse", "mean"),
            mean_physics_violation=("test_physics_violation", "mean"),
        )
        .reset_index()
        .sort_values(["hidden_dim", "data_fraction", "noise_level", "selections"], ascending=[False, False, True, False])
    )


def deduplicate_results_csv(csv_path: Path) -> None:
    """Drop duplicate completed rows using the same normalised key as resume."""
    if not csv_path.exists():
        return
    df = pd.read_csv(csv_path)
    if df.empty or not set(KEY_COLS).issubset(df.columns):
        return
    before = len(df)
    # Normal drop_duplicates is not enough for float round-trip variants
    # (0.08000000000000002 vs 0.08), so build a canonical helper key.
    helper = df[KEY_COLS].apply(make_df_key, axis=1)
    df = df.loc[~helper.duplicated(keep="last")].copy()
    if len(df) != before:
        df = sort_results(df)
        df.to_csv(csv_path, index=False)
        print(f"Resume cleanup: removed {before - len(df)} duplicate rows from existing CSV.")


def build_report(csv_path: Path, xlsx_path: Path) -> None:
    if not csv_path.exists():
        return
    df = pd.read_csv(csv_path)
    if df.empty:
        return

    # Keep last in case the script was interrupted/restarted.
    if set(KEY_COLS).issubset(df.columns):
        df = df.drop_duplicates(KEY_COLS, keep="last")
    df = sort_results(df)
    df.to_csv(csv_path, index=False)

    final_epoch = int(df["checkpoint_epoch"].max())
    final = df[df["checkpoint_epoch"] == final_epoch].copy()

    tables = {
        "All_Runs": df,
        "Final_Runs": final,
        "Summary_Model": aggregate_with_ci(final, ["model", "model_display"]),
        "Summary_Config": aggregate_with_ci(final, ["model", "model_display", "hidden_dim", "data_fraction", "noise_level", "physics_lambda"]),
        "Summary_Hidden": aggregate_with_ci(final, ["model", "model_display", "hidden_dim"]),
        "Summary_DataFraction": aggregate_with_ci(final, ["model", "model_display", "data_fraction"]),
        "Summary_Noise": aggregate_with_ci(final, ["model", "model_display", "noise_level"]),
        "Summary_Lambda": aggregate_with_ci(final[final["model"] == "TINY_GNN_PINN"], ["physics_lambda", "hidden_dim", "data_fraction", "noise_level"]),
        "Lambda_Selected_By_Val": build_lambda_selection_table(df),
        "Paired_TTests": build_paired_ttests(df),
    }

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        readme = pd.DataFrame([
            ["Generated at", datetime.now().isoformat(timespec="seconds")],
            ["Rows", len(df)],
            ["Final checkpoint", final_epoch],
            ["Physics mismatch", "data generated with data_alpha and reaction_mu; PINN uses pure diffusion with pinn_alpha"],
            ["Generalisation split", "train/validation/test are independent trajectories"],
            ["Data fraction", "subsamples complete training trajectories"],
            ["Model selection", "lowest validation MSE; selected_epoch is reported"],
            ["Inference timing", "mean/std over inference_repeats repeated forward passes"],
            ["Statistics", "Summary sheets include 95% CI; Paired_TTests has paired t-tests"],
        ], columns=["Field", "Value"])
        readme.to_excel(writer, sheet_name="README", index=False)
        for name, table in tables.items():
            table.to_excel(writer, sheet_name=name[:31], index=False)

    # Light formatting without making the file generation fragile.
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        wb = load_workbook(xlsx_path)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.sheet_view.showGridLines = False
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                width = min(max(len(str(ws.cell(1, col_idx).value or "")) + 2, 10), 34)
                ws.column_dimensions[letter].width = width
            if ws.max_row >= 2 and ws.max_column >= 1:
                ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                try:
                    table_name = "tbl_" + "".join(ch for ch in ws.title if ch.isalnum())[:20]
                    tab = Table(displayName=table_name, ref=ref)
                    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
                    ws.add_table(tab)
                except Exception:
                    ws.auto_filter.ref = ref
        wb.save(xlsx_path)
    except Exception:
        pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Corrected TFG TinyGNN/FullGNN sweep.")
    parser.add_argument("--mode", choices=["smoke", "recommended", "full"], default="recommended")
    parser.add_argument("--epochs", default=",".join(map(str, DEFAULT_CHECKPOINTS)), help="Comma list, e.g. 50,100,300")
    parser.add_argument("--seeds", default="40:50", help="Comma list or inclusive range, e.g. 40:50")
    parser.add_argument("--hidden-dims", default=",".join(map(str, DEFAULT_HIDDEN_DIMS)))
    parser.add_argument("--data-fractions", default=",".join(map(str, DEFAULT_DATA_FRACTIONS)))
    parser.add_argument("--noise-levels", default=",".join(map(str, DEFAULT_NOISE_LEVELS)))
    parser.add_argument("--physics-lambdas", default=",".join(map(str, DEFAULT_PHYSICS_LAMBDAS)))
    parser.add_argument("--learning-rate", type=float, default=1e-3)
    parser.add_argument("--dropout", type=float, default=0.1, help="Only used by FullGNN")
    parser.add_argument("--device", default="auto", help="auto, cuda, cuda:0, cpu")
    parser.add_argument("--output-dir", default=str(CONFIG.sweep_results_dir / "corrected_experiments"))
    parser.add_argument("--only-models", default="", help="Optional comma list: MLP,FULL_GNN,TINY_GNN,TINY_GNN_PINN")

    # Corrected data generation.
    parser.add_argument("--grid-size", type=int, default=GRID_SIZE)
    parser.add_argument("--trajectory-timesteps", type=int, default=60)
    parser.add_argument("--train-trajectories", type=int, default=12)
    parser.add_argument("--val-trajectories", type=int, default=4)
    parser.add_argument("--test-trajectories", type=int, default=4)
    parser.add_argument("--data-alpha", type=float, default=ALPHA)
    parser.add_argument("--pinn-alpha", type=float, default=ALPHA * 0.8, help="Physics coefficient assumed by PINN/residual")
    parser.add_argument("--reaction-mu", type=float, default=0.02, help="Reaction term present in data but absent in PINN")
    parser.add_argument("--dt", type=float, default=DT)
    parser.add_argument("--dx", type=float, default=DX)

    # Early stopping / robust timing.
    parser.add_argument("--early-stopping", action=argparse.BooleanOptionalAction, default=True)
    parser.add_argument("--min-epochs", type=int, default=50)
    parser.add_argument("--patience", type=int, default=40)
    parser.add_argument("--min-delta", type=float, default=1e-8)
    parser.add_argument("--inference-repeats", type=int, default=1000)
    parser.add_argument("--inference-warmup", type=int, default=50)

    parser.add_argument("--limit-trainings", type=int, default=0, help="Debug: stop after N model trainings. 0 = no limit.")
    parser.add_argument("--no-excel", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    started_at = time.perf_counter()

    if args.mode == "smoke":
        checkpoints = [5]
        seeds = [40]
        hidden_dims = [16]
        data_fractions = [1.0]
        noise_levels = [0.0]
        physics_lambdas = [0.1]
        args.trajectory_timesteps = min(args.trajectory_timesteps, 20)
        args.train_trajectories = min(args.train_trajectories, 3)
        args.val_trajectories = min(args.val_trajectories, 1)
        args.test_trajectories = min(args.test_trajectories, 1)
        args.inference_repeats = min(args.inference_repeats, 20)
        args.inference_warmup = min(args.inference_warmup, 5)
        args.min_epochs = 1
        args.patience = 5
    else:
        checkpoints = sorted(unique_preserve_order(parse_int_list(args.epochs)))
        seeds = parse_int_list(args.seeds)
        hidden_dims = parse_int_list(args.hidden_dims)
        data_fractions = parse_float_list(args.data_fractions)
        noise_levels = parse_float_list(args.noise_levels)
        physics_lambdas = parse_float_list(args.physics_lambdas)

    only_models = [x.strip().upper() for x in args.only_models.split(",") if x.strip()]
    device = resolve_device(args.device)
    configure_torch(device)

    combos = make_sweep_combos(args.mode, hidden_dims, data_fractions, noise_levels)
    specs = make_model_specs(only_models, physics_lambdas)
    output_dir = Path(args.output_dir)
    csv_path = output_dir / "all_results_corrected.csv"
    xlsx_path = output_dir / "sweep_report_corrected.xlsx"
    output_dir.mkdir(parents=True, exist_ok=True)

    trainings = len(seeds) * len(combos) * len(specs)
    print("\n" + "=" * 92)
    print("TFG corrected experiment sweep")
    print("=" * 92)
    print(f"Mode:                    {args.mode}")
    print(f"Device:                  {device}")
    print(f"Checkpoints:             {checkpoints}")
    print(f"Seeds:                   {seeds}")
    print(f"Combos h/data/noise:      {len(combos)}")
    print(f"Model trainings:          {trainings}")
    print(f"Trajectory split:         train={args.train_trajectories}, val={args.val_trajectories}, test={args.test_trajectories}")
    print(f"Trajectory timesteps:     {args.trajectory_timesteps}")
    print(f"Physics mismatch:         data_alpha={args.data_alpha}, pinn_alpha={args.pinn_alpha}, reaction_mu={args.reaction_mu}")
    print(f"Early stopping:           {args.early_stopping}, patience={args.patience}, min_epochs={args.min_epochs}")
    print(f"Inference repeats:        {args.inference_repeats}")
    print(f"CSV:                      {csv_path}")
    print(f"Excel:                    {xlsx_path}")
    print("=" * 92 + "\n")

    completed = load_completed_keys(csv_path)
    if completed:
        # Clean any duplicates left by interrupted/restarted runs before planning resume.
        deduplicate_results_csv(csv_path)
        completed = load_completed_keys(csv_path)
        print(f"Resume: found {len(completed)} completed rows.\n")

    training_counter = 0
    for seed in seeds:
        for combo_idx, combo in enumerate(combos, start=1):
            print(
                f"[seed {seed}] combo {combo_idx}/{len(combos)} | "
                f"hidden={combo.hidden_dim}, data_fraction={combo.data_fraction}, noise={combo.noise_level}"
            )
            data = prepare_data(seed=seed, combo=combo, args=args, device=device)
            print(
                f"  data: train_traj={data.train_trajectories_kept}/{data.train_trajectories_total}, "
                f"pairs={data.train_pairs}, val_pairs={data.val_pairs}, test_pairs={data.test_pairs}"
            )

            for spec in specs:
                if expected_already_done(
                    completed=completed,
                    mode=args.mode,
                    seed=seed,
                    checkpoints=checkpoints,
                    combo=combo,
                    spec=spec,
                    args=args,
                ):
                    print(f"  SKIP {spec.display_name:15s} lambda={spec.physics_lambda:g} (already done)")
                    continue

                training_counter += 1
                if args.limit_trainings and training_counter > args.limit_trainings:
                    print("Reached --limit-trainings; stopping early.")
                    if not args.no_excel:
                        build_report(csv_path, xlsx_path)
                    return

                print(f"  TRAIN {spec.display_name:15s} lambda={spec.physics_lambda:g}")
                rows = train_model_with_selection(
                    mode=args.mode,
                    spec=spec,
                    combo=combo,
                    seed=seed,
                    data=data,
                    checkpoints=checkpoints,
                    args=args,
                    device=device,
                )
                rows_to_write = []
                for row in rows:
                    key = make_row_key(row)
                    if key not in completed:
                        rows_to_write.append(row)
                        completed.add(key)
                append_rows_to_csv(rows_to_write, csv_path)
                for row in rows_to_write:
                    print(
                        f"    checkpoint={row['checkpoint_epoch']:>3d} | selected={row['selected_epoch']:>3d} | "
                        f"val={row['val_mse']:.6e} | test={row['test_mse']:.6e} | "
                        f"phys={row['test_physics_violation']:.6e} | "
                        f"infer={row['inference_time_ms']:.4f}±{row['inference_time_std_ms']:.4f} ms"
                    )

            del data
            if device.startswith("cuda"):
                torch.cuda.empty_cache()

    if not args.no_excel:
        print("\nBuilding Excel report with confidence intervals and t-tests...")
        build_report(csv_path, xlsx_path)

    elapsed = time.perf_counter() - started_at
    print("\nDone.")
    print(f"Elapsed: {elapsed:.1f} s")
    print(f"CSV:   {csv_path}")
    if not args.no_excel:
        print(f"Excel: {xlsx_path}")


if __name__ == "__main__":
    main()
